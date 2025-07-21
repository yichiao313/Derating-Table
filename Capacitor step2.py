from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 載入 Excel
wb = load_workbook("step2_capacitors.xlsx")
ws = wb.active  # 預設只有一個工作表

# 找出 A 欄要合併的區段（從第2列開始，跳過標題）
merge_start = 2
while merge_start <= ws.max_row:
    current_value = ws[f"A{merge_start}"].value
    merge_end = merge_start

    # 找出有相同值的連續列
    while merge_end + 1 <= ws.max_row and ws[f"A{merge_end + 1}"].value == current_value:
        merge_end += 1

    if merge_end > merge_start:
        ws.merge_cells(start_row=merge_start, start_column=1, end_row=merge_end, end_column=1)

    merge_start = merge_end + 1

# 儲存新的 Excel
wb.save("step2_capacitors_merged.xlsx")
print("✅ 已合併完成並另存為 step2_capacitors_merged.xlsx")
